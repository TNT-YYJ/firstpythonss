## 将多个excel中的内容合并在一个文件中 xlrd xlwt
## 一个excel有多个sheet 也需要合并



# workbook
#worksheet
#row colw cell



import openpyxl

wb=openpyxl.load_workbook('E:\\example.xlsx')

# Getting sheets from the workbook
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)
#创建
mysheet=wb.create_sheet('mysheet')
##单独获取
sheet3=wb.get_sheet_by_name('3')
she4=wb['mysheet']

## 单元格：
ws=wb.active
print(ws['A1'])
print(ws['A1'].value)
a=ws['A1']
print('row {},column is {}'.format(a.row,a.column,a.value))
print(a.coordinate)
ws.cell(row=1,column=2)##  从1 开始的

for i in range(1,8,2):
    print(i,ws.cell(row=i,column=1).value)

colC=ws['C']# 列
row6=ws[6]# 行
colrange=ws['B:C']
rowrange=ws[2:8]

# for col in colrange:
#     for cell in col:
#         print(cell.value)
#
# for row in ws.iter_rows(min_row=1,max_row=2):
#     for cell in row:
#         print(cell)
## 元祖==不能改的数组而已 0 开始
tuple1=tuple(ws.rows)
print('元祖',(tuple1[0] )[0].value )

# cell_range=ws['A1:C3']
# for rowofObj in cell_range:
#     for cellObj in rowofObj:
#         print(cellObj.coordinate,cellObj.value)

# from openpyxl.utils import get_column_letter,column_index_from_string
# print('get_column_letter \t ',get_column_letter(2))
# print('column_index_from_string \t ',column_index_from_string('B'))




